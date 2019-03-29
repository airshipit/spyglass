# Copyright 2018 AT&T Intellectual Property.  All other rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import logging
import pprint
import re
import sys

from openpyxl import load_workbook
from openpyxl import Workbook
import yaml

from spyglass.data_extractor.custom_exceptions import NoSpecMatched

LOG = logging.getLogger(__name__)


class ExcelParser(object):
    """Parse data from excel into a dict"""

    def __init__(self, file_name, excel_specs):
        self.file_name = file_name
        with open(excel_specs, "r") as f:
            spec_raw_data = f.read()
        self.excel_specs = yaml.safe_load(spec_raw_data)
        # A combined design spec, returns a workbook object after combining
        # all the inputs excel specs
        combined_design_spec = self.combine_excel_design_specs(file_name)
        self.wb_combined = combined_design_spec
        self.filenames = file_name
        self.spec = "xl_spec"

    @staticmethod
    def sanitize(string):
        """Remove extra spaces and convert string to lower case"""

        return string.replace(" ", "").lower()

    def compare(self, string1, string2):
        """Compare the strings"""

        return bool(re.search(self.sanitize(string1), self.sanitize(string2)))

    def validate_sheet(self, spec, sheet):
        """Check if the sheet is correct or not"""

        ws = self.wb_combined[sheet]
        header_row = self.excel_specs["specs"][spec]["header_row"]
        ipmi_header = self.excel_specs["specs"][spec]["ipmi_address_header"]
        ipmi_column = self.excel_specs["specs"][spec]["ipmi_address_col"]
        header_value = ws.cell(row=header_row, column=ipmi_column).value
        return bool(self.compare(ipmi_header, header_value))

    def find_correct_spec(self):
        """Find the correct spec"""

        for spec in self.excel_specs["specs"]:
            sheet_name = self.excel_specs["specs"][spec]["ipmi_sheet_name"]
            for sheet in self.wb_combined.sheetnames:
                if self.compare(sheet_name, sheet):
                    self.excel_specs["specs"][spec]["ipmi_sheet_name"] = sheet
                    if self.validate_sheet(spec, sheet):
                        return spec
        raise NoSpecMatched(self.excel_specs)

    def get_ipmi_data(self):
        """Read IPMI data from the sheet"""

        ipmi_data = {}
        hosts = []
        spec_ = self.excel_specs["specs"][self.spec]
        provided_sheetname = spec_["ipmi_sheet_name"]
        workbook_object, extracted_sheetname = \
            self.get_xl_obj_and_sheetname(provided_sheetname)
        if workbook_object is not None:
            ws = workbook_object[extracted_sheetname]
        else:
            ws = self.wb_combined[provided_sheetname]
        row = spec_["start_row"]
        end_row = spec_["end_row"]
        hostname_col = spec_["hostname_col"]
        ipmi_address_col = spec_["ipmi_address_col"]
        host_profile_col = spec_["host_profile_col"]
        ipmi_gateway_col = spec_["ipmi_gateway_col"]
        previous_server_gateway = None
        while row <= end_row:
            hostname = \
                self.sanitize(ws.cell(row=row, column=hostname_col).value)
            hosts.append(hostname)
            ipmi_address = ws.cell(row=row, column=ipmi_address_col).value
            if "/" in ipmi_address:
                ipmi_address = ipmi_address.split("/")[0]
            ipmi_gateway = ws.cell(row=row, column=ipmi_gateway_col).value
            if ipmi_gateway:
                previous_server_gateway = ipmi_gateway
            else:
                ipmi_gateway = previous_server_gateway
            host_profile = ws.cell(row=row, column=host_profile_col).value
            try:
                if host_profile is None:
                    raise RuntimeError("No value read from "
                                       "{} sheet:{} row:{}, col:{}".format(
                                           self.file_name, self.spec, row,
                                           host_profile_col))
            except RuntimeError as rerror:
                LOG.critical(rerror)
                sys.exit("Tugboat exited!!")
            ipmi_data[hostname] = {
                "ipmi_address": ipmi_address,
                "ipmi_gateway": ipmi_gateway,
                "host_profile": host_profile,
                "type": type,  # FIXME (Ian Pittwood): shadows type built-in
            }
            row += 1
        LOG.debug("ipmi data extracted from excel:\n{}".format(
            pprint.pformat(ipmi_data)))
        LOG.debug("host data extracted from excel:\n{}".format(
            pprint.pformat(hosts)))
        return [ipmi_data, hosts]

    def get_private_vlan_data(self, ws):
        """Get private vlan data from private IP sheet"""

        vlan_data = {}
        row = self.excel_specs["specs"][self.spec]["vlan_start_row"]
        end_row = self.excel_specs["specs"][self.spec]["vlan_end_row"]
        type_col = self.excel_specs["specs"][self.spec]["net_type_col"]
        vlan_col = self.excel_specs["specs"][self.spec]["vlan_col"]
        while row <= end_row:
            cell_value = ws.cell(row=row, column=type_col).value
            if cell_value:
                vlan = ws.cell(row=row, column=vlan_col).value
                if vlan:
                    vlan = vlan.lower()
                vlan_data[vlan] = cell_value
            row += 1
        LOG.debug("vlan data extracted from excel:\n%s" %
                  pprint.pformat(vlan_data))
        return vlan_data

    def get_private_network_data(self):
        """Read network data from the private ip sheet"""

        spec_ = self.excel_specs["specs"][self.spec]
        provided_sheetname = spec_["private_ip_sheet"]
        workbook_object, extracted_sheetname = \
            self.get_xl_obj_and_sheetname(provided_sheetname)
        if workbook_object is not None:
            ws = workbook_object[extracted_sheetname]
        else:
            ws = self.wb_combined[provided_sheetname]
        vlan_data = self.get_private_vlan_data(ws)
        network_data = {}
        row = spec_["net_start_row"]
        end_row = spec_["net_end_row"]
        col = spec_["net_col"]
        vlan_col = spec_["net_vlan_col"]
        old_vlan = ""
        while row <= end_row:
            vlan = ws.cell(row=row, column=vlan_col).value
            if vlan:
                vlan = vlan.lower()
            network = ws.cell(row=row, column=col).value
            if vlan and network:
                net_type = vlan_data[vlan]
                if "vlan" not in network_data:
                    network_data[net_type] = {"vlan": vlan, "subnet": []}
            elif not vlan and network:
                # If vlan is not present then assign old vlan to vlan as vlan
                # value is spread over several rows
                vlan = old_vlan
            else:
                row += 1
                continue
            network_data[vlan_data[vlan]]["subnet"].append(network)
            old_vlan = vlan
            row += 1
        for network in network_data:
            network_data[network]["is_common"] = True
            """
            if len(network_data[network]['subnet']) > 1:
                network_data[network]['is_common'] = False
            else:
                network_data[network]['is_common'] = True
        LOG.debug("private network data extracted from excel:\n%s"
                  % pprint.pformat(network_data))
            """
        return network_data

    def get_public_network_data(self):
        """Read public network data from public ip data"""

        spec_ = self.excel_specs["specs"][self.spec]
        provided_sheetname = spec_["public_ip_sheet"]
        workbook_object, extracted_sheetname = self.get_xl_obj_and_sheetname(
            provided_sheetname)
        if workbook_object is not None:
            ws = workbook_object[extracted_sheetname]
        else:
            ws = self.wb_combined[provided_sheetname]
        oam_row = spec_["oam_ip_row"]
        oam_col = spec_["oam_ip_col"]
        oam_vlan_col = spec_["oam_vlan_col"]
        ingress_row = spec_["ingress_ip_row"]
        oob_row = spec_["oob_net_row"]
        col = spec_["oob_net_start_col"]
        end_col = spec_["oob_net_end_col"]
        network_data = {
            "oam": {
                "subnet": [ws.cell(row=oam_row, column=oam_col).value],
                "vlan": ws.cell(row=oam_row, column=oam_vlan_col).value,
            },
            "ingress": ws.cell(row=ingress_row, column=oam_col).value,
            "oob": {
                "subnet": [],
            }
        }
        while col <= end_col:
            cell_value = ws.cell(row=oob_row, column=col).value
            if cell_value:
                network_data["oob"]["subnet"].append(self.sanitize(cell_value))
            col += 1
        LOG.debug("public network data extracted from excel:\n%s" %
                  pprint.pformat(network_data))
        return network_data

    def get_site_info(self):
        """Read location, dns, ntp and ldap data"""

        spec_ = self.excel_specs["specs"][self.spec]
        provided_sheetname = spec_["dns_ntp_ldap_sheet"]
        workbook_object, extracted_sheetname = \
            self.get_xl_obj_and_sheetname(provided_sheetname)
        if workbook_object is not None:
            ws = workbook_object[extracted_sheetname]
        else:
            ws = self.wb_combined[provided_sheetname]
        dns_row = spec_["dns_row"]
        dns_col = spec_["dns_col"]
        ntp_row = spec_["ntp_row"]
        ntp_col = spec_["ntp_col"]
        domain_row = spec_["domain_row"]
        domain_col = spec_["domain_col"]
        login_domain_row = spec_["login_domain_row"]
        ldap_col = spec_["ldap_col"]
        global_group = spec_["global_group"]
        ldap_search_url_row = spec_["ldap_search_url_row"]
        dns_servers = ws.cell(row=dns_row, column=dns_col).value
        ntp_servers = ws.cell(row=ntp_row, column=ntp_col).value
        try:
            if dns_servers is None:
                raise RuntimeError("No value for dns_server from: "
                                   "{} Sheet:'{}' Row:{} Col:{}".format(
                                       self.file_name, provided_sheetname,
                                       dns_row, dns_col))
            if ntp_servers is None:
                raise RuntimeError("No value for ntp_server from: "
                                   "{} Sheet:'{}' Row:{} Col:{}".format(
                                       self.file_name, provided_sheetname,
                                       ntp_row, ntp_col))
        except RuntimeError as rerror:
            LOG.critical(rerror)
            sys.exit("Tugboat exited!!")

        dns_servers = dns_servers.replace("\n", " ")
        ntp_servers = ntp_servers.replace("\n", " ")
        if "," in dns_servers:
            dns_servers = dns_servers.split(",")
        else:
            dns_servers = dns_servers.split()
        if "," in ntp_servers:
            ntp_servers = ntp_servers.split(",")
        else:
            ntp_servers = ntp_servers.split()
        site_info = {
            "location": self.get_location_data(),
            "dns": dns_servers,
            "ntp": ntp_servers,
            "domain": ws.cell(row=domain_row, column=domain_col).value,
            "ldap": {
                "subdomain": ws.cell(row=login_domain_row,
                                     column=ldap_col).value,
                "common_name": ws.cell(row=global_group,
                                       column=ldap_col).value,
                "url": ws.cell(row=ldap_search_url_row, column=ldap_col).value,
            },
        }
        LOG.debug(
            "Site Info extracted from\
                          excel:\n%s",
            pprint.pformat(site_info),
        )
        return site_info

    def get_location_data(self):
        """Read location data from the site and zone sheet"""

        spec_ = self.excel_specs["specs"][self.spec]
        provided_sheetname = spec_["location_sheet"]
        workbook_object, extracted_sheetname = \
            self.get_xl_obj_and_sheetname(provided_sheetname)
        if workbook_object is not None:
            ws = workbook_object[extracted_sheetname]
        else:
            ws = self.wb_combined[provided_sheetname]
        corridor_row = spec_["corridor_row"]
        column = spec_["column"]
        site_name_row = spec_["site_name_row"]
        state_name_row = spec_["state_name_row"]
        country_name_row = spec_["country_name_row"]
        clli_name_row = spec_["clli_name_row"]
        return {
            "corridor": ws.cell(row=corridor_row, column=column).value,
            "name": ws.cell(row=site_name_row, column=column).value,
            "state": ws.cell(row=state_name_row, column=column).value,
            "country": ws.cell(row=country_name_row, column=column).value,
            "physical_location": ws.cell(row=clli_name_row,
                                         column=column).value,
        }

    def validate_sheet_names_with_spec(self):
        """Checks is sheet name in spec file matches with excel file"""

        spec = list(self.excel_specs["specs"].keys())[0]
        spec_item = self.excel_specs["specs"][spec]
        sheet_name_list = []
        ipmi_header_sheet_name = spec_item["ipmi_sheet_name"]
        sheet_name_list.append(ipmi_header_sheet_name)
        private_ip_sheet_name = spec_item["private_ip_sheet"]
        sheet_name_list.append(private_ip_sheet_name)
        public_ip_sheet_name = spec_item["public_ip_sheet"]
        sheet_name_list.append(public_ip_sheet_name)
        dns_ntp_ldap_sheet_name = spec_item["dns_ntp_ldap_sheet"]
        sheet_name_list.append(dns_ntp_ldap_sheet_name)
        location_sheet_name = spec_item["location_sheet"]
        sheet_name_list.append(location_sheet_name)
        try:
            for sheetname in sheet_name_list:
                workbook_object, extracted_sheetname = \
                    self.get_xl_obj_and_sheetname(sheetname)
                if workbook_object is not None:
                    wb = workbook_object
                    sheetname = extracted_sheetname
                else:
                    wb = self.wb_combined

                if sheetname not in wb.sheetnames:
                    raise RuntimeError(
                        "SheetName '{}' not found ".format(sheetname))
        except RuntimeError as rerror:
            LOG.critical(rerror)
            sys.exit("Tugboat exited!!")

        LOG.info("Sheet names in excel spec validated")

    def get_data(self):
        """Create a dict with combined data"""

        self.validate_sheet_names_with_spec()
        ipmi_data = self.get_ipmi_data()
        network_data = self.get_private_network_data()
        public_network_data = self.get_public_network_data()
        site_info_data = self.get_site_info()
        data = {
            "ipmi_data": ipmi_data,
            "network_data": {
                "private": network_data,
                "public": public_network_data,
            },
            "site_info": site_info_data,
        }
        LOG.debug("Location data extracted from excel:\n%s" %
                  pprint.pformat(data))
        return data

    def combine_excel_design_specs(self, filenames):
        """Combines multiple excel file to a single design spec"""

        design_spec = Workbook()
        for exel_file in filenames:
            loaded_workbook = load_workbook(exel_file, data_only=True)
            for names in loaded_workbook.sheetnames:
                design_spec_worksheet = design_spec.create_sheet(names)
                loaded_workbook_ws = loaded_workbook[names]
                for row in loaded_workbook_ws:
                    for cell in row:
                        design_spec_worksheet[cell.coordinate].value = \
                            cell.value
        return design_spec

    def get_xl_obj_and_sheetname(self, sheetname):
        """The logic confirms if the sheetname is specified for example as:

        'MTN57a_AEC_Network_Design_v1.6.xlsx:Public IPs'
        """

        if re.search(".xlsx", sheetname) or re.search(".xls", sheetname):
            # Extract file name
            source_xl_file = sheetname.split(":")[0]
            wb = load_workbook(source_xl_file, data_only=True)
            return [wb, sheetname.split(":")[1]]
        else:
            return [None, sheetname]
